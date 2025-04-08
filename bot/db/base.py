import logging
import os
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from ..config import logPath


class ExcelDB:
    """
    Класс для работы с Excel-файлом как с базой данных.
    Предоставляет функции для добавления, обновления, удаления и получения данных.

    :param path: Путь к файлу Excel.
    :param sheet_name: Имя листа в Excel (по умолчанию "Sheet1").
    :param headers: Список заголовков колонок.
    :param id_column: Имя колонки, которая используется как идентификатор (по умолчанию "id").
    """

    def __init__(
        self,
        path: str,
        sheet_name: str = "Sheet1",
        id_column: str = "id",
        headers: List[str] = ["name"],
    ) -> None:
        """
        Инициализация ExcelDB, загрузка существующего файла или создание нового.

        :param path: Путь к Excel файлу
        :param sheet_name: Имя листа
        :param headers: Заголовки для колонок
        :param id_column: Имя колонки с уникальным идентификатором
        """

        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler(logPath + '/excel_db.log')
        handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        self.logger.addHandler(handler)

        self.path = path
        self.sheet_name = sheet_name
        self.id_column = id_column

        headers.insert(0, id_column)
        self.headers = headers

        self.logger.info(f"Инициализация ExcelDB для файла: {self.path}")

        if os.path.exists(self.path):
            self.wb = openpyxl.load_workbook(self.path)
            self.ws = self.wb[self.sheet_name]
            self.logger.info(f"Загружен существующий файл: {self.path}")
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = self.sheet_name
            self._create_header()
            self.save()
            self.logger.info(f"Создан новый файл и инициализированы заголовки: {self.headers}")

    def _create_header(self) -> None:
        """
        Создание заголовков для таблицы Excel, используя переданный список заголовков.
        """
        for col_num, header in enumerate(self.headers, start=1):
            self.ws[get_column_letter(col_num) + "1"] = header
        self.logger.debug("Заголовки созданы в таблице.")

    def save(self) -> None:
        """
        Сохраняет изменения в Excel файл.
        """
        self.wb.save(self.path)
        self.logger.info(f"Изменения сохранены в файл {self.path}")

    def get_all(self) -> List[tuple]:
        """
        Получить все данные из таблицы.

        :return: Список кортежей с данными из всех строк (кроме заголовков)
        """
        rows = [tuple(row) for row in self.ws.iter_rows(min_row=2, values_only=True)]
        self.logger.debug(f"Получены все записи: {rows}")
        return rows

    def get_by_id(self, record_id: int) -> Optional[tuple]:
        """
        Получить данные по уникальному идентификатору.

        :param record_id: Идентификатор записи
        :return: Кортеж с данными или None, если запись не найдена
        """
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[0] == record_id:
                self.logger.info(f"Запись с ID {record_id} найдена: {row}")
                return row
        self.logger.warning(f"Запись с ID {record_id} не найдена.")
        return None

    def add(self, data: Dict[str, Any]) -> int:
        """
        Добавить новую запись в таблицу.

        :param data: Словарь с данными для добавления, в котором обязательно должен быть 'id'
        :return: Идентификатор добавленной записи
        """
        record_id = data.get("id")

        if record_id is None:
            raise ValueError("ID не может быть пустым!")

        if self.get_by_id(record_id):
            raise ValueError(f"Запись с ID {record_id} уже существует!")

        row = [data.get(header) for header in self.headers]
        self.ws.append(row)
        self.save()
        self.logger.info(f"Добавлена новая запись с ID {data['id']}")
        return data["id"]

    def update(self, record_id: int, updates: Dict[str, Any]) -> bool:
        """
        Обновить данные записи по уникальному идентификатору.

        :param record_id: Идентификатор записи
        :param updates: Словарь с обновлениями
        :return: True, если обновление прошло успешно, иначе False
        """
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
            if row[0].value == record_id:
                for key, value in updates.items():
                    col_idx = self._get_column_index(key)
                    if col_idx:
                        self.ws.cell(row=row[0].row, column=col_idx, value=value)
                self.save()
                self.logger.info(f"Обновлена запись с ID {record_id} с новыми данными: {updates}")
                return True
        self.logger.warning(f"Запись с ID {record_id} не найдена для обновления.")
        return False

    def delete(self, record_id: int) -> bool:
        """
        Удалить запись по уникальному идентификатору.

        :param record_id: Идентификатор записи
        :return: True, если запись удалена, иначе False
        """
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
            if row[0].value == record_id:
                self.ws.delete_rows(row[0].row)
                self.save()
                self.logger.info(f"Удалена запись с ID {record_id}")
                return True
        self.logger.warning(f"Запись с ID {record_id} не найдена для удаления.")
        return False

    def _get_column_index(self, column_name: str) -> Optional[int]:
        """
        Получить индекс колонки по её имени.

        :param column_name: Имя колонки
        :return: Индекс колонки (1-based) или None, если колонка не найдена
        """
        if column_name in self.headers:
            return self.headers.index(column_name) + 1
        self.logger.warning(f"Колонка {column_name} не найдена.")
        return None
